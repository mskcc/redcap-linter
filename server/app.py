import logging
import json
from datetime import datetime
from collections import OrderedDict

import flask
from flask import request
import pandas as pd
from fuzzywuzzy import fuzz
from flask_cors import CORS, cross_origin
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from models.redcap_field import RedcapField
from api.redcap.redcap_api import RedcapApi
from linter import linter
from utils import utils
from serializers import serializer
from api.resolve import RESOLVE
from api.export import EXPORT

app = flask.Flask(__name__)
app.register_blueprint(RESOLVE)
app.register_blueprint(EXPORT)

@app.route('/save_fields', methods=['GET', 'POST', 'OPTIONS'])
def save_fields():
    form = request.form.to_dict()
    data_field_to_redcap_field_map = json.loads(form.get('dataFieldToRedcapFieldMap'))
    csv_headers = json.loads(form.get('csvHeaders'))
    existing_records = json.loads(form.get('existingRecords'))
    recordid_field = json.loads(form.get('recordidField'))
    project_info = json.loads(form.get('projectInfo'))
    malformed_sheets = json.loads(form.get('malformedSheets') or '""')
    token = json.loads(form.get('token'))
    env = json.loads(form.get('env'))
    json_data = json.loads(form.get('jsonData'), object_pairs_hook=OrderedDict)
    records = {}
    for sheet in json_data:
        matched_field_dict = data_field_to_redcap_field_map.get(sheet, {})
        csv_headers[sheet] = [matched_field_dict.get(c) or c for c in csv_headers[sheet] if matched_field_dict.get(c) != '']
        frame = pd.DataFrame(json_data[sheet])
        frame.fillna('', inplace=True)
        frame = frame.rename(index=str, columns=matched_field_dict)
        frame = frame[csv_headers[sheet]]
        records[sheet] = frame

    dd_data = json.loads(form.get('ddData'))
    dd = [RedcapField.from_json(field) for field in dd_data]

    if token:
        redcap_api = RedcapApi(env)
        if project_info['record_autonumbering_enabled'] == 1:
            if not project_info['secondary_unique_field']:
                existing_records = None
            else:
                secondary_unique_field_values = set()
                for sheet_name, sheet in records.items():
                    for _, record in sheet.iterrows():
                        if record.get(project_info['secondary_unique_field']):
                            secondary_unique_field_values.add(utils.get_record_id(record.get(project_info['secondary_unique_field'])))
                secondary_unique_field_values = list(secondary_unique_field_values)
                options = {
                    'secondary_unique_field': project_info['secondary_unique_field'],
                    'secondary_unique_field_values': secondary_unique_field_values
                }
                existing_records = redcap_api.export_records(token, options)
        else:
            record_ids = set()
            for sheet_name, sheet in records.items():
                for _, record in sheet.iterrows():
                    if record.get(recordid_field):
                        record_ids.add(utils.get_record_id(record.get(recordid_field)))
            record_ids = list(record_ids)
            options = {'records': record_ids}
            existing_records = redcap_api.export_records(token, options)

    datafile_errors = linter.lint_datafile(dd, records, project_info)
    cells_with_errors = datafile_errors['cells_with_errors']
    rows_in_error = datafile_errors['rows_in_error']
    encoded_records = datafile_errors['encoded_records']

    columns_in_error = utils.get_columns_with_errors(cells_with_errors, records)

    all_errors = [{"Error": error} for error in datafile_errors['linting_errors']]

    json_data = {}
    output_records = {}

    # logging.warning(encoded_records)
    for sheet_name, sheet in records.items():
        json_data[sheet_name] = json.loads(sheet.to_json(orient='records', date_format='iso'))
        cells_with_errors[sheet_name] = json.loads(cells_with_errors[sheet_name].to_json(orient='records'))

    for sheet_name in encoded_records:
        if malformed_sheets and sheet_name in malformed_sheets:
            continue
        output_records[sheet_name] = json.loads(encoded_records[sheet_name].to_json(orient='records'))

    records_to_reconcile = {}
    if existing_records:
        for record in existing_records:
            if not records_to_reconcile.get(record[recordid_field]):
                records_to_reconcile[record[recordid_field]] = []
            if record.get(recordid_field):
                records_to_reconcile[record[recordid_field]].append(record)

    decoded_records = {}

    for sheet_name in records.keys():
        sheet = records.get(sheet_name)
        for _, row in sheet.iterrows():
            if row.get(recordid_field):
                recordid = row[recordid_field]
                if isinstance(row.get(recordid_field), float) and row.get(recordid_field).is_integer():
                    # Excel reads this in as a float :/
                    recordid = str(int(row[recordid_field]))
                if records_to_reconcile.get(recordid):
                    decoded_rows = serializer.decode_sheet(dd, project_info, records_to_reconcile.get(recordid))
                    decoded_records[recordid] = decoded_rows

    results = {
        'jsonData':                   json_data,
        'rowsInError':                rows_in_error,
        'cellsWithErrors':            cells_with_errors,
        'allErrors':                  all_errors,
        'csvHeaders':                 csv_headers,
        'recordsToReconcile':         records_to_reconcile,
        'existingRecords':            existing_records,
        'columnsInError':             columns_in_error,
        'encodedRecords':             output_records,
        'decodedRecords':             decoded_records,
        'fieldsSaved':                True,
    }
    response = flask.jsonify(results)
    response.headers.add('Access-Control-Allow-Origin', '*')
    return response

@app.route('/import_records', methods=['GET', 'POST', 'OPTIONS'])
def import_records():
    form = request.form.to_dict()
    encoded_records = json.loads(form.get('encodedRecords'))

    env = json.loads(form.get('env'))
    token = json.loads(form.get('token'))
    redcap_api = RedcapApi(env)

    errors = {}
    for sheet_name, sheet_records in encoded_records.items():
        response = redcap_api.import_records(token, sheet_records)
        errors[sheet_name] = response

    results = {
        'importErrors': errors
    }
    response = flask.jsonify(results)
    response.headers.add('Access-Control-Allow-Origin', '*')
    return response

@app.route('/', methods=['GET', 'POST', 'OPTIONS'])
def post_form():
    records = pd.read_excel(request.files['dataFile'], sheet_name=None)
    date_cols = []
    records_with_format = load_workbook(request.files['dataFile'])
    for sheet in records_with_format.sheetnames:
        for row in records_with_format[sheet].iter_rows(min_row=2):
            for cell in row:
                # MRN
                column_letter = get_column_letter(cell.column)
                column_header = records_with_format[sheet][column_letter + '1'].value
                if column_header in records[sheet].columns and cell.number_format == '00000000':
                    current_list = list(records[sheet][column_header])
                    current_list = [str(i).rjust(8, '0') if isinstance(i, int) else i for i in current_list]
                    records[sheet][column_header] = current_list
                if column_header in records[sheet].columns and cell.number_format == 'mm-dd-yy':
                    date_cols.append(column_header)
                    current_list = list(records[sheet][column_header])
                    current_list = [i.strftime('%m/%d/%Y') if isinstance(i, datetime) and not pd.isnull(i) else i for i in current_list]
                    records[sheet][column_header] = current_list
            break

    form = request.form.to_dict()
    token = form.get('token')
    env = form.get('environment')
    datafile_name = form.get('dataFileName')
    mappings = None
    existing_records = None
    form_names = set()
    form_name_to_dd_fields = {}
    data_field_to_redcap_field_map = {}
    data_field_to_choice_map = {}
    original_to_correct_value_map = {}
    no_match_redcap_fields = []

    if 'mappingsFile' in request.files:
        mappings = pd.read_excel(request.files['mappingsFile'], sheet_name="Sheet1")

        if list(mappings["dataFieldToRedcapFieldMap"]):
            data_field_to_redcap_field_map = json.loads(list(mappings["dataFieldToRedcapFieldMap"])[0])
        if list(mappings["dataFieldToChoiceMap"]):
            data_field_to_choice_map = json.loads(list(mappings["dataFieldToChoiceMap"])[0])
        if list(mappings["originalToCorrectedValueMap"]):
            original_to_correct_value_map = json.loads(list(mappings["originalToCorrectedValueMap"])[0])
        if list(mappings["noMatchRedcapFields"]):
            no_match_redcap_fields = json.loads(list(mappings["noMatchRedcapFields"])[0])

    redcap_api = RedcapApi(env)

    project_info = {
        'custom_record_label': '',
        'secondary_unique_field': '',
        'record_autonumbering_enabled': 0,
        'repeatable_instruments': [],
        'next_record_name': 1
    }

    data_dictionary = None
    existing_records = None
    if token:
        try:
            data_dictionary = redcap_api.fetch_data_dictionary(token)
            project_info = redcap_api.fetch_project_info(token)
            project_info['next_record_name'] = redcap_api.generate_next_record_name(token)
            if project_info['has_repeating_instruments_or_events'] == 1:
                repeatable_instruments = redcap_api.fetch_repeatable_instruments(token)
                project_info['repeatable_instruments'] = [i['form_name'] for i in repeatable_instruments]
            if project_info['record_autonumbering_enabled'] == 0:
                data_dictionary[0]['required'] = 'Y'
            dd = [RedcapField.from_json(field) for field in data_dictionary]
        except Exception as e:
            logging.warning(e)
            results = {'error': "Error: {0}".format(e)}
            response = flask.jsonify(results)
            response.headers.add('Access-Control-Allow-Origin', '*')
            return response
    else:
        data_dictionary_name = form.get('dataDictionaryName')
        if data_dictionary_name.endswith('.csv'):
            dd_df = pd.read_csv(request.files['dataDictionary'])
            dd_df.fillna('', inplace=True)
        elif data_dictionary_name.endswith('.xlsx') or data_dictionary_name.endswith('.xls'):
            dd_df = pd.read_excel(request.files['dataDictionary'])
        dd = [RedcapField.from_data_dictionary(dd_df, field) for field in list(dd_df['Variable / Field Name'])]
        if dd[0].field_name == 'record_id':
            project_info['record_autonumbering_enabled'] = 1
        if 'existingRecordsFile' in request.files:
            existing_records = pd.read_csv(request.files['existingRecordsFile'])
            existing_records = json.loads(existing_records.to_json(orient='records', date_format='iso'))

    all_csv_headers = []
    dd_headers = []
    dd_data = {}
    dd_data_raw = {}
    if data_dictionary is not None:
        dd_headers = list(data_dictionary[0].keys())
        dd_data_raw = data_dictionary
    else:
        dd_headers = list(dd_df.columns)
        dd_data_raw = json.loads(dd_df.to_json(orient='records', date_format='iso'))

    dd_data = [field.__dict__ for field in dd]

    for dd_field in dd:
        if not form_name_to_dd_fields.get(dd_field.form_name):
            form_name_to_dd_fields[dd_field.form_name] = []
        form_name_to_dd_fields.get(dd_field.form_name).append(dd_field.field_name)
        form_names.add(dd_field.form_name)

    recordid_field = 'record_id'
    if not project_info['record_autonumbering_enabled']:
        recordid_field = dd[0].field_name

    form_names = list(form_names)

    for sheet_name, sheet in records.items():
        all_csv_headers += list(sheet.columns)
        all_csv_headers = [i for i in all_csv_headers if 'Unnamed' not in i]

    all_field_names = [f.field_name for f in dd]

    redcap_field_candidates = {}
    data_field_candidates = {}
    csv_headers = {}
    fields_not_in_redcap = {}
    duplicate_fields = {}

    for sheet_name, sheet in records.items():
        duplicate_fields[sheet_name] = {}
        # Remove empty rows
        sheet.dropna(axis=0, how='all', inplace=True)
        csv_headers[sheet_name] = list(sheet.columns)
        csv_headers[sheet_name] = [item for item in csv_headers[sheet_name] if 'Unnamed' not in item]
        for header in csv_headers[sheet_name]:
            duplicate_fields[sheet_name][header] = duplicate_fields[sheet_name].get(header, 0) + 1
        duplicate_fields[sheet_name] = [k for k, v in duplicate_fields[sheet_name].items() if v > 1]
        normalized_headers = utils.parameterize_list(csv_headers[sheet_name])
        fields_not_in_redcap[sheet_name] = [header for header, normalized_header in zip(csv_headers[sheet_name], normalized_headers) if normalized_header not in all_field_names]

    all_csv_headers = list(set(all_csv_headers))

    unmatched_data_fields = {}

    for sheet in csv_headers:
        data_field_to_redcap_field_map[sheet] = data_field_to_redcap_field_map.get(sheet, {})
        unmatched_data_fields[sheet] = unmatched_data_fields.get(sheet, [])
        for header in csv_headers[sheet]:
            normalized_header = utils.parameterize(header)
            if data_field_to_redcap_field_map[sheet].get(header):
                continue
            if normalized_header in all_field_names:
                data_field_to_redcap_field_map[sheet][header] = normalized_header
            else:
                unmatched_data_fields[sheet].append(header)

    selected_columns = {}

    matched_redcap_fields = []
    matched_redcap_fields += no_match_redcap_fields
    for sheet_name, field_map in data_field_to_redcap_field_map.items():
        selected_columns[sheet_name] = field_map.keys()
        matched_redcap_fields += field_map.values()
    unmatched_redcap_fields = [f for f in all_field_names if f not in matched_redcap_fields and f != 'record_id']
    for f1 in all_field_names:
        dd_field = [f for f in dd_data if f['field_name'] == f1][0]
        for sheet in csv_headers:
            for f2 in csv_headers[sheet]:
                if not redcap_field_candidates.get(f1):
                    redcap_field_candidates[f1] = []
                redcap_field_candidates[f1].append({
                    'candidate': f2,
                    'sheets': [sheet],
                    'score': max(fuzz.token_set_ratio(f1, f2), fuzz.token_set_ratio(dd_field['field_label'], f2))
                })

    for sheet in csv_headers:
        for f1 in csv_headers[sheet]:
            if data_field_candidates.get(f1):
                continue
            data_field_candidates[f1] = []
            for f2 in all_field_names:
                dd_field = [f for f in dd_data if f['field_name'] == f2][0]
                data_field_candidates[f1].append({
                    'candidate': f2,
                    'form_name': dd_field['form_name'],
                    'score': max(fuzz.token_set_ratio(f1, f2), fuzz.token_set_ratio(dd_field['field_label'], f1))
                })

    malformed_sheets = []
    all_errors = []

    form_names = [redcap_field.form_name for redcap_field in dd]
    form_names = list(set(form_names))
    for sheet_name in records.keys():
        sheet = records.get(sheet_name)

        redcap_field_names = [f.field_name for f in dd]

        matching_fields = [f for f in sheet.columns if f in redcap_field_names]
        if not matching_fields and not data_field_to_redcap_field_map.get(sheet_name):
            malformed_sheets.append(sheet_name)

        redcap_fields_not_in_data = [f for f in redcap_field_names if f not in sheet.columns]

        sheet_fields_not_in_redcap = [f for f in sheet.columns if f not in redcap_field_names]

        if sheet_fields_not_in_redcap:
            all_errors.append("Fields in Instrument {0} not present in REDCap: {1}".format(sheet_name, str(sheet_fields_not_in_redcap)))
        if redcap_fields_not_in_data:
            all_errors.append("Fields in REDCap not present in Instrument {0}: {1}".format(sheet_name, str(redcap_fields_not_in_data)))

    json_data = {}

    for sheet_name, sheet in records.items():
        json_data[sheet_name] = json.loads(sheet.to_json(orient='records', date_format='iso'))

    results = {
        'csvHeaders':              csv_headers,
        'jsonData':                json_data,
        'ddHeaders':               dd_headers,
        'ddData':                  dd_data,
        'ddDataRaw':               dd_data_raw,
        'formNames':               form_names,
        'dateColumns':             date_cols,
        'duplicateFields':         duplicate_fields,
        'malformedSheets':         malformed_sheets,
        'recordFieldsNotInRedcap': fields_not_in_redcap,
        'formNameToDdFields':      form_name_to_dd_fields,
        'projectInfo':             project_info,
        'existingRecords':         existing_records,
        'recordidField':           recordid_field,
        'redcapFieldCandidates':   redcap_field_candidates,
        'dataFieldCandidates':     data_field_candidates,
        'unmatchedRedcapFields':   unmatched_redcap_fields,
        'unmatchedDataFields':     unmatched_data_fields,
        'dataFileName':            datafile_name,
        'token':                   token,
        'env':                     env,
    }
    if data_field_to_redcap_field_map:
        results['dataFieldToRedcapFieldMap'] = data_field_to_redcap_field_map
    if data_field_to_choice_map:
        results['dataFieldToChoiceMap'] = data_field_to_choice_map
    if original_to_correct_value_map:
        results['originalToCorrectedValueMap'] = original_to_correct_value_map
    if no_match_redcap_fields:
        results['noMatchRedcapFields'] = no_match_redcap_fields

    response = flask.jsonify(results)
    response.headers.add('Access-Control-Allow-Origin', '*')
    return response

CORS(app, expose_headers='Authorization')
